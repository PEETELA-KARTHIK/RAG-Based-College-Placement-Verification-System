"""
Structured Excel Parser for RAG.
Extracts each row as a separate chunk with rich metadata to correct off-by-one errors.
"""

from openpyxl import load_workbook
from datetime import datetime, timezone
from config import CHUNK_SIZE

def clean_key(key: str) -> str:
    """Normalize header keys (lowercase, snake_case)."""
    if not key:
        return "unknown_col"
    return str(key).strip().lower().replace(" ", "_").replace(".", "").replace("/", "_")

def parse_excel(file_path: str, document_name: str, category: str) -> list[dict]:
    """
    Parse Excel file row-by-row into structured chunks.
    
    Returns:
        List of dicts: {
            "text": "Name: John | ID: 123 ...", 
            "metadata": {
                "document_name": ...,
                "category": ...,
                "sheet_name": ...,
                "row_number": ...,
                "name": ...,  # Extracted if present
                "roll_no": ..., # Extracted if present
                # ... other fields
            }
        }
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    chunks = []
    upload_date = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        
        if not rows:
            continue
            
        # Headers (Row 1)
        raw_headers = rows[0]
        # Create a map of index -> cleaned_key
        header_map = {}
        for idx, h in enumerate(raw_headers):
            if h is not None:
                header_map[idx] = clean_key(h)
                
        # Iterate data rows
        for i, row in enumerate(rows[1:], start=2): # 1-based index, skip header
            
            # 1. Build Metadata & Text
            row_metadata = {
                "document_name": document_name,
                "category": category,
                "upload_date": upload_date,
                "sheet_name": sheet_name,
                "row_number": i,
                "source_type": "excel_row" # Maker for verification logic
            }
            
            text_parts = []
            has_content = False
            
            for idx, value in enumerate(row):
                if idx in header_map and value is not None:
                    str_val = str(value).strip()
                    if str_val:
                        key = header_map[idx]
                        # Add to metadata (store as string for Chroma compatibility)
                        row_metadata[key] = str_val
                        # Add to text representation
                        text_parts.append(f"{key}: {str_val}")
                        has_content = True
            
            if not has_content:
                continue
                
            # context text for embedding/LLM
            # We prefix with identifying info to help retrieval
            row_text = f"Record in {document_name} ({sheet_name}, Row {i}): " + " | ".join(text_parts)
            
            chunks.append({
                "text": row_text,
                "metadata": row_metadata
            })
            
    wb.close()
    return chunks
